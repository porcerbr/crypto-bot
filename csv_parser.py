"""
csv_parser.py — Parser para CSVs do Investing.com (formato PT-BR).

Suporta:
  - Números brasileiros: "1.234,56" → 1234.56
  - Datas: "25/04/2026" → datetime
  - Colunas: Data, Preço (close), Abertura (open), Alta (high), Baixa (low)
  - Linhas de resumo ao final são ignoradas automaticamente
  - Barras em ordem cronológica (mais antiga primeiro)
  - Arquivos Excel renomeados para .csv (detecção automática)
"""

from __future__ import annotations

import csv
import io
from datetime import datetime, timezone, date
from typing import Any


def _br_float(s: Any) -> float | None:
    """'1.234,56' → 1234.56 | '0,8634' → 0.8634 | aceita números puros."""
    if s is None:
        return None
    if isinstance(s, (int, float)):
        try:
            return float(s)
        except (TypeError, ValueError):
            return None
    try:
        txt = str(s).strip()
        if not txt:
            return None
        return float(txt.replace('.', '').replace(',', '.'))
    except (ValueError, AttributeError):
        return None


def _parse_date(s: Any) -> datetime | None:
    """Converte data textual ou datetime/ date para datetime UTC."""
    if s is None:
        return None
    if isinstance(s, datetime):
        return s if s.tzinfo else s.replace(tzinfo=timezone.utc)
    if isinstance(s, date):
        return datetime(s.year, s.month, s.day, tzinfo=timezone.utc)

    txt = str(s).strip()
    if not txt:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(txt, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def _normalize_rows(rows: list[list[Any]]) -> list[dict]:
    """Normaliza linhas brutas (CSV ou XLSX) para OHLC."""
    if not rows:
        return []

    header = [str(h).strip().lower().replace('"', '') if h is not None else '' for h in rows[0]]
    col_map = {
        "data": "date",
        "date": "date",
        "time": "date",
        "preço": "close",
        "preco": "close",
        "price": "close",
        "fechamento": "close",
        "abertura": "open",
        "open": "open",
        "alta": "high",
        "high": "high",
        "máxima": "high",
        "maxima": "high",
        "baixa": "low",
        "low": "low",
        "mínima": "low",
        "minima": "low",
    }

    idx: dict[str, int] = {}
    for i, h in enumerate(header):
        key = col_map.get(h)
        if key and key not in idx:
            idx[key] = i

    required = {"date", "close", "open", "high", "low"}
    if not required.issubset(idx):
        # fallback: tenta descobrir olhando a primeira linha útil
        idx = {"date": 0, "close": 1, "open": 2, "high": 3, "low": 4}

    bars: list[dict] = []
    for row in rows[1:]:
        if not row or len(row) < 5:
            continue

        def safe_get(key: str):
            i = idx.get(key)
            return row[i] if i is not None and i < len(row) else None

        raw_date = safe_get("date")
        if raw_date is None:
            continue
        if isinstance(raw_date, str):
            raw_date = raw_date.strip().strip('"')
            if not raw_date or ':' in raw_date:
                continue

        ts = _parse_date(raw_date)
        close = _br_float(safe_get("close"))
        open_ = _br_float(safe_get("open"))
        high = _br_float(safe_get("high"))
        low = _br_float(safe_get("low"))

        if not all([ts, close, open_, high, low]):
            continue
        if not (low <= close <= high and low <= open_ <= high):
            continue

        bars.append({
            "timestamp": ts,
            "open": open_,
            "high": high,
            "low": low,
            "close": close,
        })

    bars.sort(key=lambda b: b["timestamp"])
    return bars


def _parse_xlsx_bytes(content: bytes) -> list[dict]:
    """Tenta ler planilha Excel renomeada para CSV."""
    try:
        from openpyxl import load_workbook
    except Exception:
        return []

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if not rows:
            return []

        # Alguns arquivos ".csv" vindos do Excel chegam como uma única coluna
        # contendo a linha inteira em texto CSV. Nesse caso, reconstrói o texto.
        single_cell_rows = sum(1 for r in rows if len(r) == 1)
        if single_cell_rows >= max(1, int(len(rows) * 0.8)):
            text_lines: list[str] = []
            for r in rows:
                if not r:
                    continue
                cell = r[0]
                if cell is None:
                    continue
                text_lines.append(str(cell))
            if text_lines:
                reader = csv.reader(io.StringIO("\n".join(text_lines), newline=''))
                return _normalize_rows(list(reader))

        return _normalize_rows(rows)
    except Exception:
        return []


def parse_investing_csv(content: str | bytes) -> list[dict]:
    """
    Parseia conteúdo CSV do Investing.com PT-BR.

    Também aceita arquivos XLSX renomeados para .csv.

    Retorna lista de dicts ordenada do mais antigo para o mais recente:
      [{"timestamp": datetime, "open": float, "high": float,
        "low": float, "close": float}, ...]
    """
    if isinstance(content, bytes):
        # Detecta arquivo Excel renomeado (ZIP/XLSX)
        if content[:2] == b'PK':
            parsed = _parse_xlsx_bytes(content)
            if parsed:
                return parsed

        # Tenta UTF-8 com BOM, depois latin-1 (comum em CSVs do Windows/Excel)
        for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
            try:
                content = content.decode(enc, errors="strict")
                break
            except (UnicodeDecodeError, AttributeError):
                continue
        else:
            content = content.decode("utf-8-sig", errors="replace")

    content = content.replace("\r\n", "\n").replace("\r", "\n")
    reader = csv.reader(io.StringIO(content, newline=''))
    rows = list(reader)
    return _normalize_rows(rows)


# Autodetecta símbolo pelo range de preços (heurística)
_PRICE_RANGES = [
    ("EURUSD", 0.90, 1.30),
    ("GBPUSD", 1.10, 1.60),
    ("USDCAD", 1.20, 1.50),
    ("USDCHF", 0.75, 1.10),
    ("EURGBP", 0.78, 0.96),
    ("AUDUSD", 0.55, 0.85),
    ("NZDUSD", 0.50, 0.76),
    ("USDJPY", 95.0, 170.0),
    ("EURJPY", 110.0, 190.0),
    ("GBPJPY", 140.0, 225.0),
    ("XAUUSD", 1200.0, 3800.0),
]

def detect_symbol(bars: list[dict], filename: str = "") -> str | None:
    """
    Tenta identificar o símbolo.
    1. Tenta pelo nome do arquivo (ex: AUDUSD.csv, eurusd_weekly.csv)
    2. Usa o preço mediano de TODAS as barras (não apenas as 10 primeiras)
       para evitar falso positivo em períodos de volatilidade.
    """
    if not bars:
        return None

    fname_up = filename.upper()
    for sym, _, _ in _PRICE_RANGES:
        if sym in fname_up:
            return sym

    closes = sorted(b["close"] for b in bars)
    median = closes[len(closes) // 2]
    sorted_ranges = sorted(_PRICE_RANGES, key=lambda x: x[2] - x[1])
    for sym, lo, hi in sorted_ranges:
        if lo <= median <= hi:
            return sym
    return None
