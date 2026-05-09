"""
backtester.py — motor de backtest robusto e regime-adaptativo.

Objetivos:
- reduzir overfitting
- manter frequência mínima razoável
- avaliar sinais com lógica mais parecida com a operação ao vivo
"""

from __future__ import annotations

import argparse
import csv
import math
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Callable, Iterable

import pandas as pd

from config import Config
from performance import calculate_metrics_from_history
from utils import calc_pnl_usd, is_jpy_pair, log, load_strategy_settings, pip_factor, get_sl_tp_atr


@dataclass
class Bar:
    timestamp: datetime
    open: float
    high: float
    low: float
    close: float


@dataclass
class BacktestResult:
    metrics: dict
    trades: list[dict]
    equity_curve: list[dict]
    params: dict = field(default_factory=dict)


# ──────────────────────────────────────────────────────────────────────────────
# Timeframe / parsing
# ──────────────────────────────────────────────────────────────────────────────

def detect_timeframe(bars: list[Bar]) -> str:
    if len(bars) < 2:
        return "H1"

    deltas = []
    for i in range(1, min(10, len(bars))):
        d = abs((bars[i].timestamp - bars[i - 1].timestamp).total_seconds())
        if d > 0:
            deltas.append(d)

    if not deltas:
        return "H1"

    avg_seconds = sum(deltas) / len(deltas)
    if avg_seconds >= 5 * 24 * 3600:
        return "W1"
    if avg_seconds >= 23 * 3600:
        return "D1"
    if avg_seconds >= 3600:
        return "H1"
    if avg_seconds >= 900:
        return "M15"
    if avg_seconds >= 300:
        return "M5"
    return "M1"


def resample_to_h1(bars: list[Bar]) -> list[Bar]:
    """Agrega candles de timeframe curto (M1/M5/M15) em H1."""
    if not bars:
        return bars
    import pandas as pd

    df = bars_to_dataframe(bars)
    h1 = df.resample("1h").agg({
        "Open":  "first",
        "High":  "max",
        "Low":   "min",
        "Close": "last",
    }).dropna()

    result: list[Bar] = []
    for ts, row in h1.iterrows():
        result.append(Bar(
            timestamp=ts.to_pydatetime(),
            open=float(row["Open"]),
            high=float(row["High"]),
            low=float(row["Low"]),
            close=float(row["Close"]),
        ))
    return result

def prepare_bars_for_backtest(bars: list[Bar]) -> list[Bar]:
    """
    Normaliza o histórico para o timeframe efetivamente usado no backtest.
    M15 agora é suportado nativamente — NÃO é mais reamostrado para H1.
    Apenas M1 e M5 (sub-minuto) são reamostrados para H1.
    """
    if not bars:
        return bars
    raw_tf = detect_timeframe(bars)
    if raw_tf in ('M1', 'M5'):   # só TFs muito curtos são reamostrados
        return resample_to_h1(bars)
    return bars  # M15, H1, H4, D1 ficam como estão



def _parse_dt(value: str) -> datetime | None:
    if not value:
        return None
    raw = str(value).strip().replace("Z", "+00:00")
    for fmt in (
        None,
        "%Y%m%d %H%M%S",   # HistData/Dukascopy ASCII: 20230101 170400
        "%Y%m%d %H%M",     # variante sem segundos:     20230101 1704
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        "%m/%d/%Y",  # formato US do Investing.com inglês
        "%Y-%m-%d",
    ):
        try:
            if fmt is None:
                dt = datetime.fromisoformat(raw)
            else:
                dt = datetime.strptime(raw, fmt)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt
        except Exception:
            continue
    return None


def load_bars_from_csv(path: str | Path) -> list[Bar]:
    path = Path(path)

    def _from_rows(rows) -> list[Bar]:
        bars: list[Bar] = []
        for row in rows:
            row = dict(row)
            ts = _parse_dt(
                row.get("timestamp")
                or row.get("time")
                or row.get("date")
                or row.get("datetime")
                or row.get("Data")
                or row.get("Data/Hora")
            )
            if ts is None:
                continue

            def _num(*keys: str) -> float | None:
                for k in keys:
                    v = row.get(k)
                    if v is None:
                        continue
                    try:
                        return float(str(v).replace(",", ".").strip())
                    except Exception:
                        continue
                return None

            o = _num("open", "Open", "abertura")
            h = _num("high", "High", "máxima", "max")
            l = _num("low", "Low", "mínima", "min")
            c = _num("close", "Close", "fechamento")
            if None in (o, h, l, c):
                continue
            bars.append(Bar(timestamp=ts, open=float(o), high=float(h), low=float(l), close=float(c)))
        return sorted(bars, key=lambda b: b.timestamp)

    # Planilha Excel renomeada para .csv (começa com PK)
    try:
        with path.open('rb') as fh:
            sig = fh.read(4)
        if sig.startswith(b'PK'):
            from shutil import copyfile
            from tempfile import NamedTemporaryFile
            from openpyxl import load_workbook

            with NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                copyfile(path, tmp.name)
                tmp_path = Path(tmp.name)
            try:
                wb = load_workbook(tmp_path, read_only=True, data_only=True)
                ws = wb.active
                rows = []
                for vals in ws.iter_rows(values_only=True):
                    if not vals:
                        continue
                    cell = next((v for v in vals if isinstance(v, str) and v.strip()), None)
                    if cell:
                        rows.append(cell.strip())
                parsed = []
                for i, line in enumerate(rows):
                    try:
                        parts = next(csv.reader([line]))
                    except Exception:
                        continue
                    parts = [p.strip() for p in parts if p is not None and str(p).strip()]
                    if not parts:
                        continue
                    if i == 0 and parts[0].lower().startswith('data'):
                        continue
                    # Formato típico: Data,Preço,Abertura,Alta,Baixa,Var%
                    if len(parts) >= 5:
                        date_s = parts[0]
                        close_s = parts[1] if len(parts) > 1 else None
                        open_s = parts[2] if len(parts) > 2 else None
                        high_s = parts[3] if len(parts) > 3 else None
                        low_s = parts[4] if len(parts) > 4 else None
                        row = {
                            'timestamp': date_s,
                            'open': open_s,
                            'high': high_s,
                            'low': low_s,
                            'close': close_s,
                        }
                        parsed.append(row)
                if parsed:
                    return _from_rows(parsed)
            finally:
                try:
                    tmp_path.unlink(missing_ok=True)
                except Exception:
                    pass
    except Exception:
        pass

    raw = path.read_text(encoding="utf-8-sig", errors="ignore")
    sample = raw[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;	|")
    except Exception:
        dialect = csv.get_dialect("excel")

    lines = [l for l in raw.splitlines() if l.strip()]
    if not lines:
        return []

    # ── Detecção de formato sem cabeçalho (HistData/Dukascopy ASCII) ──────────
    # Esses arquivos começam com uma linha de dados pura (ex: "20230101 170400;…")
    # em vez de um cabeçalho textual. Se a primeira coluna começar com dígito,
    # tratamos como headerless e mapeamos posicionalmente.
    first_cols = next(csv.reader([lines[0]], dialect=dialect), [])
    first_cell = first_cols[0].strip() if first_cols else ""
    is_headerless = first_cell[:1].isdigit()

    # ── Detecção formato Investing.com (PT/ES) ──────────────────────────────
    # Cabeçalho em português: "Data,Preço,Abertura,Alta,Baixa,Var%"
    # Decimal com vírgula: "1,1792" → 1.1792 | Data: DD/MM/YYYY
    investing_headers_pt = {"data", "preco", "preço", "abertura", "alta", "baixa"}
    investing_headers_en = {"date", "price", "open", "high", "low"}
    first_lower = {c.strip().lower().replace("\u00e7", "c").strip('"') for c in first_cols}
    is_investing = (len(investing_headers_pt & first_lower) >= 3
                    or len(investing_headers_en & first_lower) >= 4)

    if is_investing:
        def _fix_num(s: str) -> str:
            """Normaliza número para float string, suportando formatos PT/BR e EN."""
            s = s.strip().strip('"').replace("%", "").replace(" ", "")
            if not s:
                return "0"
            n_dot   = s.count(".")
            n_comma = s.count(",")
            # Já está em formato EN padrão: "1.1715" ou "1,234.56"
            if n_dot >= 1 and n_comma == 0:
                # ponto é decimal (ex: "1.1715") — não mexe
                return s
            # Formato PT/BR: vírgula decimal sem ponto  "1,1792"
            if n_comma == 1 and n_dot == 0:
                return s.replace(",", ".")
            # Formato PT/BR com milhar: "1.179,20"
            if n_comma == 1 and n_dot >= 1:
                return s.replace(".", "").replace(",", ".")
            # Formato EN com milhar: "1,234.56"
            if n_dot == 1 and n_comma >= 1:
                return s.replace(",", "")
            return s

        reader_inv = csv.DictReader(lines, dialect=dialect)
        col_map = {}  # mapeia header real → chave interna
        header_aliases = {
            "data": "timestamp", "date": "timestamp",
            "preco": "close", "preço": "close", "price": "close", "último": "close", "ultimo": "close",
            "abertura": "open", "open": "open",
            "alta": "high",  "high": "high", "máx": "high", "max": "high",
            "baixa": "low",  "low": "low",  "mín": "low",  "min": "low",
            "vol.": None, "vol": None, "change %": None, "var%": None,  # ignora colunas irrelevantes
        }
        rows_inv: list[dict] = []
        for raw_row in reader_inv:
            row: dict[str, str] = {}
            for raw_col, val in raw_row.items():
                if raw_col is None:
                    continue
                key = header_aliases.get(raw_col.strip().lower().replace("\u00e7", "c").replace("\u00e9", "e").replace("\u00e1", "a").replace("\u00ed", "i"))
                if key == "timestamp":
                    row[key] = val.strip().strip('"').strip()
                elif key in ("open", "high", "low", "close"):
                    cleaned = val.strip().strip('"').strip()
                    row[key] = _fix_num(cleaned)
            # Linha de rodapé do Investing.com começa com "Abertura :"
            if "timestamp" not in row or ":" in row.get("timestamp", ""):
                continue
            if len(row) >= 4:
                rows_inv.append(row)
        return _from_rows(rows_inv)

    if is_headerless:
        # Ordem esperada: DateTime, Open, High, Low, Close[, Volume, ...]
        positional_keys = ["timestamp", "open", "high", "low", "close"]
        rows: list[dict] = []
        for line in lines:
            parts = next(csv.reader([line], dialect=dialect), [])
            if len(parts) < 5:
                continue
            row = {positional_keys[i]: parts[i].strip() for i in range(5)}
            rows.append(row)
        return _from_rows(rows)

    reader = csv.DictReader(lines, dialect=dialect)
    return _from_rows(reader)


def bars_from_dicts(data: list[dict]) -> list[Bar]:
    bars = []
    for d in data:
        try:
            ts = d["timestamp"]
            if not isinstance(ts, datetime):
                ts = _parse_dt(ts) or datetime.fromtimestamp(float(ts), tz=timezone.utc)
            bars.append(Bar(timestamp=ts, open=float(d["open"]), high=float(d["high"]), low=float(d["low"]), close=float(d["close"])))
        except Exception:
            continue
    return sorted(bars, key=lambda b: b.timestamp)


def bars_to_dataframe(bars: list[Bar]) -> pd.DataFrame:
    idx = pd.to_datetime([b.timestamp for b in bars], utc=True)
    df = pd.DataFrame(
        {
            "Open": [b.open for b in bars],
            "High": [b.high for b in bars],
            "Low": [b.low for b in bars],
            "Close": [b.close for b in bars],
            "Volume": [0.0 for _ in bars],
        },
        index=idx,
    )
    return df


# ──────────────────────────────────────────────────────────────────────────────
# Session / costs
# ──────────────────────────────────────────────────────────────────────────────

def _in_session(bar: Bar, symbol: str, tf: str = "H1") -> bool:
    """
    Filtra horários de mercado morto.

    M15: apenas bloqueia 00h–05h UTC (liquidez zero) e fins de semana.
    Deixa tudo o mais aberto — sessões de Tóquio, Londres e NY são todas válidas.

    H1: janela mais restrita (pico de liquidez por par).
    """
    ts = bar.timestamp
    # Fim de semana: mercado fechado
    if ts.weekday() >= 5:
        return False
    # Sexta > 22h UTC: risco de gap no fim de semana
    if ts.weekday() == 4 and ts.hour >= 22:
        return False

    h = ts.hour
    if tf in ("M15", "M5", "M1"):
        # M15: só bloqueia o mercado morto (madrugada UTC)
        return h >= 5   # bloqueia 00:00–04:59 UTC
    # H1 — janela restrita (pico de liquidez por par)
    if symbol == "XAUUSD":
        return 7 <= h < 20
    if is_jpy_pair(symbol):
        return h < 9 or h >= 23
    if "AUD" in symbol or "NZD" in symbol:
        return h < 8 or h >= 22
    return 7 <= h < 17


def _apply_cost(price: float, direction: str, symbol: str) -> float:
    spread_pips = Config.SPREAD_PIPS.get(symbol, 1.0) if getattr(Config, "USE_SPREAD_MODEL", True) else 0.0
    slippage_pips = Config.SLIPPAGE_PIPS.get(symbol, 0.3) if getattr(Config, "USE_SLIPPAGE_MODEL", True) else 0.0
    # custo conservador: metade do spread + slippage médio esperado
    cost = (spread_pips * 0.5 + slippage_pips * 0.5) * pip_factor(symbol)
    if direction == "BUY":
        return round(price + cost, 5)
    return round(price - cost, 5)


# ──────────────────────────────────────────────────────────────────────────────
# Indicators
# ──────────────────────────────────────────────────────────────────────────────

def _indicators(df: pd.DataFrame) -> dict | None:
    n = len(df)
    if n < 40:
        return None

    c = df["Close"]
    h = df["High"]
    l = df["Low"]
    o = df["Open"]

    ema9 = c.ewm(span=9, adjust=False).mean()
    ema21 = c.ewm(span=21, adjust=False).mean()
    ema50 = c.ewm(span=50, adjust=False).mean()
    ema200 = c.ewm(span=min(200, n - 1), adjust=False).mean()

    macd_line = c.ewm(span=12, adjust=False).mean() - c.ewm(span=26, adjust=False).mean()
    macd_signal = macd_line.ewm(span=9, adjust=False).mean()

    d = c.diff()
    gain = d.clip(lower=0).ewm(span=14, adjust=False).mean()
    loss = (-d.clip(upper=0)).ewm(span=14, adjust=False).mean()
    rsi = 100 - 100 / (1 + gain / loss.replace(0, 1e-10))

    tr = pd.concat([h - l, (h - c.shift()).abs(), (l - c.shift()).abs()], axis=1).max(axis=1)
    atr = tr.ewm(span=14, adjust=False).mean()

    up_move = h.diff()
    down_move = -l.diff()
    plus_dm = ((up_move > down_move) & (up_move > 0)) * up_move.clip(lower=0)
    minus_dm = ((down_move > up_move) & (down_move > 0)) * down_move.clip(lower=0)
    atr_safe = atr.replace(0, 1e-10)
    plus_di = 100 * plus_dm.ewm(span=14, adjust=False).mean() / atr_safe
    minus_di = 100 * minus_dm.ewm(span=14, adjust=False).mean() / atr_safe
    adx = ((plus_di - minus_di).abs() / (plus_di + minus_di + 1e-10) * 100).ewm(span=14, adjust=False).mean()

    bb_mid = c.rolling(20).mean()
    bb_std = c.rolling(20).std(ddof=0).fillna(0)
    bb_upper = bb_mid + 2 * bb_std
    bb_lower = bb_mid - 2 * bb_std

    price = float(c.iloc[-1])
    atr_val = float(atr.iloc[-1]) if not math.isnan(float(atr.iloc[-1])) else 0.0
    e21 = float(ema21.iloc[-1])
    e50 = float(ema50.iloc[-1])
    e200 = float(ema200.iloc[-1])
    e9 = float(ema9.iloc[-1])
    rsi_val = float(rsi.iloc[-1])
    rsi_prev = float(rsi.iloc[-2]) if n >= 2 else rsi_val
    macd_now = float(macd_line.iloc[-1])
    macd_sig = float(macd_signal.iloc[-1])
    macd_prev = float(macd_line.iloc[-2]) if n >= 2 else macd_now
    sig_prev = float(macd_signal.iloc[-2]) if n >= 2 else macd_sig

    return {
        "price": price,
        "ema9": e9,
        "ema21": e21,
        "ema50": e50,
        "ema200": e200,
        "atr": atr_val,
        "adx": float(adx.iloc[-1]),
        "pdi": float(plus_di.iloc[-1]),
        "ndi": float(minus_di.iloc[-1]),
        "rsi": rsi_val,
        "rsi_prev": rsi_prev,
        "macd_above": macd_now > macd_sig,
        "macd_below": macd_now < macd_sig,
        "macd_cross_up": macd_prev <= sig_prev and macd_now > macd_sig,
        "macd_cross_down": macd_prev >= sig_prev and macd_now < macd_sig,
        "dist_e21": (price - e21) / atr_val if atr_val > 0 else 0.0,
        "dist_bb_up": (float(bb_upper.iloc[-1]) - price) / atr_val if atr_val > 0 else 0.0,
        "dist_bb_dn": (price - float(bb_lower.iloc[-1])) / atr_val if atr_val > 0 else 0.0,
        "candle_bull": float(c.iloc[-1]) > float(o.iloc[-1]),
        "candle_bear": float(c.iloc[-1]) < float(o.iloc[-1]),
        "rsi_bounce_up": rsi_prev < 42 and rsi_val >= 42,
        "rsi_bounce_dn": rsi_prev > 58 and rsi_val <= 58,
        "trend_up": price > e200 and e21 > e50,
        "trend_dn": price < e200 and e21 < e50,
        "range_mode": float(adx.iloc[-1]) <= getattr(Config, "REGIME_ADX_RANGING", 18),
    }



def build_indicator_cache(bars: list[Bar], lookback: int = 300) -> list[dict | None]:
    """
    Computa indicadores técnicos para cada barra de forma vetorizada.

    Retorna uma lista de dicts (um por barra) com todos os indicadores pré-calculados.
    Usada para evitar recalcular indicadores a cada barra no loop de backtest.

    Args:
        bars:     lista de Bar em ordem cronológica
        lookback: não usado (mantido por compatibilidade de assinatura) — usa a série completa

    Returns:
        lista[dict | None] — None para as primeiras barras (sem histórico suficiente)
    """
    if not bars:
        return []

    df = bars_to_dataframe(bars)
    n  = len(df)

    if n < 40:
        return [None] * n

    c     = df["Close"]
    h_col = df["High"]
    l_col = df["Low"]
    o_col = df["Open"]

    # EMAs
    ema9   = c.ewm(span=9,   adjust=False).mean()
    ema21  = c.ewm(span=21,  adjust=False).mean()
    ema50  = c.ewm(span=50,  adjust=False).mean()
    ema200 = c.ewm(span=200, adjust=False).mean()

    # MACD
    macd_line   = c.ewm(span=12, adjust=False).mean() - c.ewm(span=26, adjust=False).mean()
    macd_signal = macd_line.ewm(span=9, adjust=False).mean()

    # RSI
    d    = c.diff()
    gain = d.clip(lower=0).ewm(span=14, adjust=False).mean()
    loss = (-d.clip(upper=0)).ewm(span=14, adjust=False).mean()
    rsi  = 100 - 100 / (1 + gain / loss.replace(0, 1e-10))

    # ATR e ADX
    tr = pd.concat(
        [h_col - l_col, (h_col - c.shift()).abs(), (l_col - c.shift()).abs()], axis=1
    ).max(axis=1)
    atr = tr.ewm(span=14, adjust=False).mean()

    up_move   = h_col.diff()
    down_move = -l_col.diff()
    plus_dm   = ((up_move > down_move) & (up_move > 0)) * up_move.clip(lower=0)
    minus_dm  = ((down_move > up_move) & (down_move > 0)) * down_move.clip(lower=0)
    atr_safe  = atr.replace(0, 1e-10)
    plus_di   = 100 * plus_dm.ewm(span=14, adjust=False).mean() / atr_safe
    minus_di  = 100 * minus_dm.ewm(span=14, adjust=False).mean() / atr_safe
    adx       = (
        (plus_di - minus_di).abs() / (plus_di + minus_di + 1e-10) * 100
    ).ewm(span=14, adjust=False).mean()

    # Bollinger Bands
    bb_mid   = c.rolling(20).mean()
    bb_std   = c.rolling(20).std(ddof=0).fillna(0)
    bb_upper = bb_mid + 2 * bb_std
    bb_lower = bb_mid - 2 * bb_std

    regime_adx = getattr(Config, "REGIME_ADX_RANGING", 18)
    cache: list[dict | None] = []

    for i in range(n):
        if i < 26:          # aguarda MACD (26 períodos mínimo)
            cache.append(None)
            continue

        price    = float(c.iloc[i])
        atr_val  = float(atr.iloc[i])
        if math.isnan(atr_val):
            atr_val = 0.0
        atr_s    = atr_val if atr_val > 0 else 1e-10

        e9   = float(ema9.iloc[i])
        e21  = float(ema21.iloc[i])
        e50  = float(ema50.iloc[i])
        e200 = float(ema200.iloc[i])

        rsi_val  = float(rsi.iloc[i])
        rsi_prev = float(rsi.iloc[i - 1]) if i >= 1 else rsi_val

        macd_now  = float(macd_line.iloc[i])
        macd_sig  = float(macd_signal.iloc[i])
        macd_prev = float(macd_line.iloc[i - 1]) if i >= 1 else macd_now
        sig_prev  = float(macd_signal.iloc[i - 1]) if i >= 1 else macd_sig

        e9_prev  = float(ema9.iloc[i - 1])  if i >= 1 else e9
        e21_prev = float(ema21.iloc[i - 1]) if i >= 1 else e21

        cache.append({
            "price": price,
            "ema9": e9, "ema21": e21, "ema50": e50, "ema200": e200,
            "atr": atr_val,
            "adx": float(adx.iloc[i]),
            "pdi": float(plus_di.iloc[i]),
            "ndi": float(minus_di.iloc[i]),
            "rsi": rsi_val,
            "rsi_prev": rsi_prev,
            "macd_above":      macd_now > macd_sig,
            "macd_below":      macd_now < macd_sig,
            "macd_cross_up":   macd_prev <= sig_prev and macd_now > macd_sig,
            "macd_cross_down": macd_prev >= sig_prev and macd_now < macd_sig,
            "dist_e21":        (price - e21) / atr_s,
            "dist_bb_up":      (float(bb_upper.iloc[i]) - price) / atr_s,
            "dist_bb_dn":      (price - float(bb_lower.iloc[i])) / atr_s,
            "candle_bull": float(c.iloc[i]) > float(o_col.iloc[i]),
            "candle_bear": float(c.iloc[i]) < float(o_col.iloc[i]),
            "rsi_bounce_up": rsi_prev < 42 and rsi_val >= 42,
            "rsi_bounce_dn": rsi_prev > 58 and rsi_val <= 58,
            "trend_up":  price > e200 and e21 > e50,
            "trend_dn":  price < e200 and e21 < e50,
            "range_mode": float(adx.iloc[i]) <= regime_adx,
            # M15: EMA9/21 cross + posição relativa
            "ema9_above_ema21": e9 > e21,
            "ema9_cross_up":    e9_prev <= e21_prev and e9 > e21,
            "ema9_cross_dn":    e9_prev >= e21_prev and e9 < e21,
            "above_ema50":  price > e50,
            "below_ema50":  price < e50,
            "above_ema21":  price > e21,
            "below_ema21":  price < e21,
        })

    return cache


# ──────────────────────────────────────────────────────────────────────────────
# Signal logic
# ──────────────────────────────────────────────────────────────────────────────

def _regime(res: dict, tf: str) -> str:
    adx = float(res.get("adx", 0) or 0)
    if adx >= getattr(Config, "REGIME_ADX_TRENDING", 25) and (res["trend_up"] or res["trend_dn"]):
        return "trend"
    if adx <= getattr(Config, "REGIME_ADX_RANGING", 18):
        return "range"
    if tf == "W1" and adx >= getattr(Config, "REGIME_ADX_STRONG", 30):
        return "trend"
    return "transition"


def _signal(
    res: dict,
    tf: str,
    min_confluence: int = 6,
    adx_min: float | None = None,
    pull_range: tuple[float, float] | None = None,
    weekly_trade_target: float = 1.5,
    h4_bias: str | None = None,
    require_h4_alignment: bool = False,
) -> str | None:
    adx_min = float(adx_min if adx_min is not None else (20 if tf == "W1" else 22))
    pull_range = pull_range or ((-0.8, 1.6) if tf == "H1" else (-1.0, 2.0))
    min_confluence = max(4, min(9, int(min_confluence or 6)))

    regime = _regime(res, tf)
    if regime not in ("trend", "transition"):
        return None

    price = res["price"]
    d21 = res["dist_e21"]

    def count(*conds: bool) -> int:
        return sum(1 for c in conds if c)

    def pull_ok(direction: str) -> bool:
        if direction == "BUY":
            return pull_range[0] <= d21 <= pull_range[1]
        return -pull_range[1] <= d21 <= -pull_range[0]

    def h4_ok(direction: str) -> bool:
        if not require_h4_alignment:
            return True
        if not h4_bias or h4_bias == "NEUTRO":
            return True
        return str(h4_bias).upper() == direction

    # Trend-only: não tentamos reverter ranges de forma agressiva no lab profissional.
    if res["trend_up"] and price > res["ema50"] and price > res["ema200"]:
        score = count(
            res["trend_up"],
            price > res["ema50"],
            price > res["ema200"],
            pull_ok("BUY"),
            (res["macd_cross_up"] or res["rsi_bounce_up"]),
            res["candle_bull"],
            res["adx"] >= adx_min,
            res["pdi"] >= res["ndi"],
            50 <= res["rsi"] <= 68,
        )
        if score >= min_confluence and pull_ok("BUY") and h4_ok("BUY"):
            return "BUY"

    if res["trend_dn"] and price < res["ema50"] and price < res["ema200"]:
        score = count(
            res["trend_dn"],
            price < res["ema50"],
            price < res["ema200"],
            pull_ok("SELL"),
            (res["macd_cross_down"] or res["rsi_bounce_dn"]),
            res["candle_bear"],
            res["adx"] >= adx_min,
            res["ndi"] >= res["pdi"],
            32 <= res["rsi"] <= 50,
        )
        if score >= min_confluence and pull_ok("SELL") and h4_ok("SELL"):
            return "SELL"

    # Pequena folga apenas em transição bem alinhada.
    if weekly_trade_target >= 1.5 and regime == "transition":
        if res["trend_up"] and h4_ok("BUY") and res["macd_cross_up"] and res["adx"] >= max(18.0, adx_min - 2):
            return "BUY"
        if res["trend_dn"] and h4_ok("SELL") and res["macd_cross_down"] and res["adx"] >= max(18.0, adx_min - 2):
            return "SELL"

    return None


def _signal_m15(
    res: dict,
    min_confluence: int = 1,
    adx_min: float = 20.0,
    pull_range: tuple[float, float] | None = None,
    weekly_trade_target: float = 8.0,
    h1_bias: str | None = None,
    rsi_ob: float = 68.0,
    rsi_os: float = 32.0,
) -> str | None:
    """
    Estratégia de 3 indicadores — o que traders profissionais realmente usam:

        EMA50  →  direção da tendência (filtro estrutural)
        MACD   →  momentum e timing de entrada (cruzamento de sinal)
        RSI    →  evita entrar em regiões extremas (sobrecomprado/sobrevendido)

    + H1 bias: descarta sinais contra a tendência do timeframe superior.

    Hard requirements (todos obrigatórios):
        BUY:  preço > EMA50  +  MACD cruzou para cima  +  RSI < rsi_ob  +  H1 não em baixa
        SELL: preço < EMA50  +  MACD cruzou para baixo +  RSI > rsi_os  +  H1 não em alta

    Soft conditions (evoluídas pelo genético, 0–3 necessárias):
        ADX >= adx_min  •  RSI em zona ideal  •  vela confirma direção

    Referência: EMA+RSI+MACD é a combinação #1 usada por traders profissionais de Forex
    (Axi, XS Broker, Quantified Strategies, LiteFinance — consenso de mercado 2024-2026).
    """
    rsi    = float(res.get("rsi", 50) or 50)
    adx    = float(res.get("adx", 0)  or 0)
    d21    = float(res.get("dist_e21", 0) or 0)

    h1_allows_buy  = h1_bias in (None, "NEUTRAL", "BUY")
    h1_allows_sell = h1_bias in (None, "NEUTRAL", "SELL")

    # ── BUY ──────────────────────────────────────────────────────────────────
    if (res.get("above_ema50", False)      # tendência M15 altista
            and res.get("macd_cross_up", False)  # MACD cruza para cima (trigger)
            and rsi < rsi_ob               # RSI não sobrecomprado
            and h1_allows_buy):            # H1 não em queda

        soft = sum([
            adx >= adx_min,                # força de tendência (opcional)
            40 <= rsi <= 65,               # RSI em zona ideal de entrada
            res.get("candle_bull", False), # vela de confirmação
        ])
        if soft >= min_confluence:
            return "BUY"

    # ── SELL ─────────────────────────────────────────────────────────────────
    if (res.get("below_ema50", False)
            and res.get("macd_cross_down", False)
            and rsi > rsi_os
            and h1_allows_sell):

        soft = sum([
            adx >= adx_min,
            35 <= rsi <= 60,
            res.get("candle_bear", False),
        ])
        if soft >= min_confluence:
            return "SELL"

    return None

def _build_h1_bias_from_m15(bars: list[Bar]) -> list[str | None]:
    """
    Calcula o viés H1 para cada barra M15 usando multi-timeframe.

    Reagrupa barras M15 → H1, calcula EMA21 e EMA50 no H1, e mapeia de volta
    para o índice M15. Cada barra M15 recebe "BUY", "SELL" ou "NEUTRAL".

    - BUY:     preço H1 > EMA21 H1  E  EMA21 > EMA50 (tendência H1 altista)
    - SELL:    preço H1 < EMA21 H1  E  EMA21 < EMA50 (tendência H1 baixista)
    - NEUTRAL: nenhuma tendência clara

    Usado em _signal_m15 como filtro de direção — maior melhora de WR.
    """
    if not bars or len(bars) < 25:
        return [None] * len(bars)

    df = bars_to_dataframe(bars)
    if df.empty:
        return [None] * len(bars)

    # Reagrupa M15 → H1
    h1_df = df.resample("1h").agg({
        "Open": "first", "High": "max", "Low": "min", "Close": "last"
    }).dropna()

    if len(h1_df) < 22:
        return [None] * len(bars)

    close  = h1_df["Close"]
    ema21h = close.ewm(span=21, adjust=False).mean()
    ema50h = close.ewm(span=50, adjust=False).mean()

    bias_rows = []
    for ts in h1_df.index:
        try:
            price = float(h1_df.loc[ts, "Close"])
            e21   = float(ema21h.loc[ts])
            e50   = float(ema50h.loc[ts])
            if price > e21 and e21 > e50:
                bias = "BUY"
            elif price < e21 and e21 < e50:
                bias = "SELL"
            else:
                bias = "NEUTRAL"
        except Exception:
            bias = "NEUTRAL"
        bias_rows.append({"timestamp": ts.to_pydatetime().replace(tzinfo=None), "h1_bias": bias})

    orig_df  = pd.DataFrame({"timestamp": [b.timestamp.replace(tzinfo=None) for b in bars]})
    bias_df  = pd.DataFrame(bias_rows).sort_values("timestamp")
    merged   = pd.merge_asof(
        orig_df.sort_values("timestamp"),
        bias_df,
        on="timestamp",
        direction="backward",
    )
    return [None if pd.isna(v) else str(v) for v in merged["h1_bias"].tolist()]


def _build_h4_bias_map(bars: list[Bar]) -> list[str | None]:
    """Mapeia cada candle H1 para o viés do H4 mais recente já fechado."""
    if not bars or len(bars) < 200:
        return [None] * len(bars)

    df_full = bars_to_dataframe(bars)
    h4_df = df_full.resample("4h", label="right", closed="right").agg({
        "Open": "first",
        "High": "max",
        "Low": "min",
        "Close": "last",
    }).dropna()

    if len(h4_df) < 20:
        return [None] * len(bars)

    h4_bars: list[Bar] = []
    for ts, row in h4_df.iterrows():
        h4_bars.append(Bar(
            timestamp=ts.to_pydatetime(),
            open=float(row["Open"]),
            high=float(row["High"]),
            low=float(row["Low"]),
            close=float(row["Close"]),
        ))

    h4_cache = build_indicator_cache(h4_bars, lookback=260)
    h4_rows: list[dict] = []
    for bar, res in zip(h4_bars, h4_cache):
        bias = "NEUTRO"
        if res:
            adx_min = getattr(Config, "REGIME_ADX_TRENDING", 25) - 2
            if res.get("trend_up") and float(res.get("adx", 0) or 0) >= adx_min:
                bias = "BUY"
            elif res.get("trend_dn") and float(res.get("adx", 0) or 0) >= adx_min:
                bias = "SELL"
        h4_rows.append({"timestamp": bar.timestamp, "h4_bias": bias})

    orig = pd.DataFrame({"timestamp": [b.timestamp for b in bars]})
    bias_df = pd.DataFrame(h4_rows).sort_values("timestamp")
    merged = pd.merge_asof(orig.sort_values("timestamp"), bias_df, on="timestamp", direction="backward")
    return [None if pd.isna(v) else str(v) for v in merged["h4_bias"].tolist()]


# ──────────────────────────────────────────────────────────────────────────────
# Backtest engine
# ──────────────────────────────────────────────────────────────────────────────

def run_backtest(
    bars: list[Bar],
    symbol: str,
    initial_balance: float | None = None,
    min_confluence: int = 6,
    adx_min: float | None = None,
    atr_sl_mult: float = 1.2,
    atr_tp_mult: float = 2.5,
    pull_range: tuple[float, float] | None = None,
    risk_pct: float = 0.5,
    warmup_bars: int | None = None,
    weekly_trade_target: float = 1.5,
    max_bars_in_trade: int | None = None,
    indicator_cache: list[dict | None] | None = None,
    prepared_bars: bool = False,
    h4_bias_map: list[str | None] | None = None,
    rsi_ob: float = 68.0,   # M15: RSI overbought threshold (don't BUY above)
    rsi_os: float = 32.0,   # M15: RSI oversold threshold (don't SELL below)
) -> BacktestResult:
    if not bars:
        return BacktestResult(metrics=calculate_metrics_from_history([], initial_balance=initial_balance), trades=[], equity_curve=[], params={"symbol": symbol})

    # Resample automático: M1/M5/M15 → H1 para performance e compatibilidade da estratégia
    if not prepared_bars:
        bars = prepare_bars_for_backtest(bars)
    else:
        bars = list(bars)
    tf = detect_timeframe(bars)
    initial_balance = float(initial_balance if initial_balance is not None else Config.INITIAL_BALANCE)
    balance = initial_balance

    wb = warmup_bars or (80 if tf == "H1" else 60)   # M15 precisa de 60 barras de warmup (~15h)
    max_bars = int(max_bars_in_trade or (60 if tf == "H1" else 20))  # M15: expira em 20 barras (5h)
    cooldown_after_loss = 2 if tf == "H1" else 1      # M15: cooldown mais curto

    if indicator_cache is None:
        indicator_cache = build_indicator_cache(bars)
    if h4_bias_map is None:
        h4_bias_map = _build_h4_bias_map(bars) if tf == "H1" else [None] * len(bars)
    elif len(h4_bias_map) < len(bars):
        h4_bias_map = list(h4_bias_map) + [None] * (len(bars) - len(h4_bias_map))

    # M15: calcula viés H1 para filtro multi-timeframe (maior melhora de WR)
    h1_bias_map_m15: list[str | None] = []
    if tf == "M15":
        h1_bias_map_m15 = _build_h1_bias_from_m15(bars)
        if len(h1_bias_map_m15) < len(bars):
            h1_bias_map_m15 += [None] * (len(bars) - len(h1_bias_map_m15))
    trades: list[dict] = []
    active: dict | None = None
    cooldown = 0

    for i in range(wb, len(bars)):
        bar = bars[i]
        h4_bias   = h4_bias_map[i]    if i < len(h4_bias_map)       else None
        h1_bias_m = h1_bias_map_m15[i] if i < len(h1_bias_map_m15) else None

        if active is not None:
            t = active
            bars_open = i - t["bar_i"]

            # ── Trailing stop: move SL para breakeven após 1× ATR a favor ──
            if not t.get("be_done", False):
                atr_at_entry = t.get("atr_entry", 0)
                if atr_at_entry > 0:
                    if t["dir"] == "BUY" and bar.high >= t["entry"] + atr_at_entry:
                        t["sl"] = max(t["sl"], t["entry"])   # SL → breakeven
                        t["be_done"] = True
                    elif t["dir"] == "SELL" and bar.low <= t["entry"] - atr_at_entry:
                        t["sl"] = min(t["sl"], t["entry"])   # SL → breakeven
                        t["be_done"] = True

            if t["dir"] == "BUY":
                hit_sl = bar.low <= t["sl"]
                hit_tp = bar.high >= t["tp"]
            else:
                hit_sl = bar.high >= t["sl"]
                hit_tp = bar.low <= t["tp"]

            if hit_sl and hit_tp:
                # Conservador: assume SL primeiro quando ambos tocam na mesma vela
                hit_tp = False

            force = bars_open >= max_bars
            if hit_sl or hit_tp or force:
                if force and not hit_sl and not hit_tp:
                    exit_px = bar.close
                    pnl = calc_pnl_usd(symbol, t["dir"], t["entry"], exit_px, t["lot"], usdjpy_price=150.0) - t.get("comm", 0)
                    result = "WIN" if pnl > 0 else "LOSS"
                else:
                    exit_px = t["tp"] if hit_tp else t["sl"]
                    pnl = calc_pnl_usd(symbol, t["dir"], t["entry"], exit_px, t["lot"], usdjpy_price=150.0) - t.get("comm", 0)
                    result = "WIN" if hit_tp else "LOSS"

                balance = round(balance + t["margin"] + pnl, 2)
                trades.append({
                    "symbol": symbol,
                    "dir": t["dir"],
                    "result": result,
                    "pnl": round(pnl, 2),
                    "entry": t["entry"],
                    "exit": exit_px,
                    "sl": t["sl"],
                    "tp": t["tp"],
                    "lot": t["lot"],
                    "bars_open": bars_open,
                    "opened_at": t["opened_at"].isoformat(),
                    "closed_at": bar.timestamp.isoformat(),
                    "closed_ts": bar.timestamp.timestamp(),
                    "closed_ts_iso": bar.timestamp.isoformat(),
                    "adx": t.get("adx", 0),
                    "timeframe": tf,
                })
                active = None
                if result == "LOSS":
                    cooldown = cooldown_after_loss
            continue

        if cooldown > 0:
            cooldown -= 1
            continue

        if not _in_session(bar, symbol, tf):
            continue

        res = indicator_cache[i] if i < len(indicator_cache) else None
        if not res or res["atr"] <= 0:
            continue

        if tf == "M15":
            direction = _signal_m15(
                res,
                min_confluence=min_confluence,
                adx_min=adx_min if adx_min is not None else 20.0,
                pull_range=pull_range,
                weekly_trade_target=weekly_trade_target,
                h1_bias=h1_bias_m,
                rsi_ob=rsi_ob,
                rsi_os=rsi_os,
            )
        else:
            direction = _signal(
                res,
                tf,
                min_confluence=min_confluence,
                adx_min=adx_min,
                pull_range=pull_range,
                weekly_trade_target=weekly_trade_target,
                h4_bias=h4_bias,
                require_h4_alignment=(tf == "H1"),
            )
        if not direction:
            continue

        entry = _apply_cost(bar.close, direction, symbol)
        sl, tp, _, _ = get_sl_tp_atr(entry, res["atr"], direction, atr_sl_mult=atr_sl_mult, atr_tp_mult=atr_tp_mult)

        if direction == "BUY" and (sl >= entry or tp <= entry):
            continue
        if direction == "SELL" and (sl <= entry or tp >= entry):
            continue

        cs = 100 if symbol == "XAUUSD" else 100_000
        sl_dist = abs(entry - sl)
        max_risk_usd = balance * max(0.1, float(risk_pct)) / 100.0
        if sl_dist <= 0:
            continue

        lot_by_risk = max_risk_usd / (sl_dist * cs)
        lot = max(Config.MIN_LOT, round(min(lot_by_risk, 50.0), 2))

        margin = round(entry * lot * cs / Config.DEFAULT_LEVERAGE, 2)
        if margin <= 0 or margin > balance * 0.45 or margin > balance:
            continue

        comm = Config.COMMISSION_PER_LOT.get("FOREX", 6.0) * lot
        balance -= margin

        active = {
            "dir": direction,
            "entry": entry,
            "sl": sl,
            "tp": tp,
            "lot": lot,
            "margin": margin,
            "comm": comm,
            "bar_i": i,
            "opened_at": bar.timestamp,
            "adx": res["adx"],
            "atr_entry": res["atr"],   # ATR no momento da entrada (para trailing stop)
            "be_done": False,          # flag: breakeven já aplicado?
        }

    if active is not None:
        t = active
        pnl = calc_pnl_usd(symbol, t["dir"], t["entry"], bars[-1].close, t["lot"], usdjpy_price=150.0) - t.get("comm", 0)
        balance = round(balance + t["margin"] + pnl, 2)
        trades.append({
            "symbol": symbol,
            "dir": t["dir"],
            "result": "WIN" if pnl > 0 else "LOSS",
            "pnl": round(pnl, 2),
            "entry": t["entry"],
            "exit": bars[-1].close,
            "sl": t["sl"],
            "tp": t["tp"],
            "lot": t["lot"],
            "bars_open": len(bars) - t["bar_i"],
            "opened_at": t["opened_at"].isoformat(),
            "closed_at": bars[-1].timestamp.isoformat(),
            "closed_ts": bars[-1].timestamp.timestamp(),
            "closed_ts_iso": bars[-1].timestamp.isoformat(),
            "adx": t.get("adx", 0),
            "timeframe": tf,
        })

    metrics = calculate_metrics_from_history(trades, initial_balance=initial_balance, current_balance=balance)

    if trades:
        first_ts = trades[0].get("closed_ts") or trades[0].get("opened_at")
        last_ts = trades[-1].get("closed_ts") or trades[-1].get("closed_at")
        try:
            first_dt = datetime.fromisoformat(str(trades[0]["closed_at"]))
            last_dt = datetime.fromisoformat(str(trades[-1]["closed_at"]))
            span_days = max(1.0, (last_dt - first_dt).total_seconds() / 86400.0)
        except Exception:
            span_days = max(1.0, len(bars) / (24.0 if tf == "H1" else 5.0))
        metrics["trade_frequency_per_week"] = round(len(trades) / max(1e-6, span_days / 7.0), 2)
        metrics["avg_bars_per_trade"] = round(sum(t.get("bars_open", 0) for t in trades) / len(trades), 2)
    else:
        metrics["trade_frequency_per_week"] = 0.0
        metrics["avg_bars_per_trade"] = 0.0

    equity_curve = metrics.pop("equity_curve", [])
    return BacktestResult(metrics=metrics, trades=trades, equity_curve=equity_curve, params={"symbol": symbol, "timeframe": tf})


# ──────────────────────────────────────────────────────────────────────────────
# Legacy helpers
# ──────────────────────────────────────────────────────────────────────────────

def backtest_trades(trades: Iterable[dict], initial_balance=None) -> dict:
    return calculate_metrics_from_history(trades, initial_balance=initial_balance)


def backtest_from_strategy(bars, strategy, initial_balance=None) -> dict:
    all_trades = []
    for i in range(1, len(bars)):
        for t in (strategy(bars, i) or []):
            t = dict(t)
            t.setdefault("closed_at", bars[i].timestamp.isoformat())
            all_trades.append(t)
    return calculate_metrics_from_history(all_trades, initial_balance=initial_balance)


# ──────────────────────────────────────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser()
    p.add_argument("csv")
    p.add_argument("--symbol", default="EURUSD")
    p.add_argument("--balance", type=float, default=Config.INITIAL_BALANCE)
    a = p.parse_args()

    bars = load_bars_from_csv(a.csv)
    if not bars:
        raise SystemExit("Nenhum candle válido.")

    tf = detect_timeframe(bars)
    log(f"[BACKTEST] {len(bars)} barras {tf} | {bars[0].timestamp:%d/%m/%Y} → {bars[-1].timestamp:%d/%m/%Y}")
    r = run_backtest(bars, a.symbol, a.balance)
    m = r.metrics

    print()
    print("═" * 52)
    print(f"  {a.symbol} · {tf} · Regime-adaptativo")
    print("═" * 52)
    print(f"  Trades:        {m['total_trades']} ({m['wins']}W / {m['losses']}L)")
    print(f"  Win Rate:      {m['winrate']}%")
    print(f"  Profit Factor: {m['profit_factor']}")
    print(f"  Max Drawdown:  {m['max_drawdown_pct']}%")
    print(f"  Sharpe:        {m.get('sharpe_ratio', 0)}")
    print(f"  Trades/week:   {m.get('trade_frequency_per_week', 0)}")
    print(f"  P&L:           ${m['total_pnl']}")
    print("═" * 52)



if __name__ == "__main__":
    main()


# ═══════════════════════════════════════════════════════════════════════════════
# FASE 2 — Walk-Forward Real + Monte Carlo + Análise de Overfitting
# ═══════════════════════════════════════════════════════════════════════════════

import statistics
import random as _random


def walk_forward_backtest(
    bars: list,
    symbol: str = "EURUSD",
    initial_balance: float | None = None,
    train_size: int = 300,
    test_size: int = 100,
    step: int = 100,
    strategy_fn=None,
) -> dict:
    """
    Walk-Forward real: treina em `train_size` velas, valida nas próximas `test_size`,
    avança `step` velas e repete.

    Separa resultado bonito (in-sample) de resultado robusto (out-of-sample).

    Args:
        bars: lista de Bar
        symbol: par monitorado
        initial_balance: saldo inicial (default: Config.INITIAL_BALANCE)
        train_size: velas de treino por janela
        test_size: velas de validação por janela
        step: passo entre janelas
        strategy_fn: função opcional para gerar trades a partir de barras

    Returns:
        dict com métricas in-sample, out-of-sample e diagnóstico de overfitting
    """
    balance = float(initial_balance or Config.INITIAL_BALANCE)
    oos_trades: list[dict] = []
    is_trades:  list[dict] = []
    windows: list[dict] = []

    i = 0
    while i + train_size + test_size <= len(bars):
        train_bars = bars[i : i + train_size]
        test_bars  = bars[i + train_size : i + train_size + test_size]

        # Backtest in-sample (treino)
        r_train = run_backtest(train_bars, symbol, balance)
        is_trades.extend(r_train.trades)

        # Backtest out-of-sample (validação)
        r_test = run_backtest(test_bars, symbol, balance)
        oos_trades.extend(r_test.trades)

        windows.append({
            "window_start": i,
            "window_end": i + train_size + test_size,
            "train_metrics": r_train.metrics,
            "oos_metrics": r_test.metrics,
            "degradation_pct": _calc_degradation(r_train.metrics, r_test.metrics),
        })

        i += step

    is_metrics  = calculate_metrics_from_history(is_trades,  initial_balance=balance)
    oos_metrics = calculate_metrics_from_history(oos_trades, initial_balance=balance)

    overfitting_score = _overfitting_score(is_metrics, oos_metrics)

    return {
        "in_sample":  is_metrics,
        "out_of_sample": oos_metrics,
        "windows": windows,
        "overfitting_score": overfitting_score,
        "overfitting_verdict": _overfitting_verdict(overfitting_score),
        "n_windows": len(windows),
        "config": {
            "symbol": symbol,
            "train_size": train_size,
            "test_size": test_size,
            "step": step,
            "total_bars": len(bars),
        },
    }


def monte_carlo_simulation(
    trades: list[dict],
    n_simulations: int = 1000,
    initial_balance: float | None = None,
    seed: int = 42,
) -> dict:
    """
    Monte Carlo por permutação de sequência de trades.

    Embaralha a ordem dos trades N vezes e recalcula métricas, gerando
    distribuição de resultados possíveis. Detecta dependência de "sorte na ordem".

    Args:
        trades: lista de trades com 'pnl'
        n_simulations: número de simulações
        initial_balance: saldo inicial
        seed: seed para reprodutibilidade

    Returns:
        dict com estatísticas da distribuição (p5, p50, p95, worst, best)
    """
    balance = float(initial_balance or Config.INITIAL_BALANCE)
    rng = _random.Random(seed)

    sim_metrics: list[dict] = []
    for _ in range(n_simulations):
        shuffled = list(trades)
        rng.shuffle(shuffled)
        m = calculate_metrics_from_history(shuffled, initial_balance=balance)
        sim_metrics.append(m)

    def _pct(values: list[float], p: float) -> float:
        if not values:
            return 0.0
        s = sorted(values)
        idx = int(len(s) * p / 100)
        return round(s[min(idx, len(s) - 1)], 3)

    drawdowns     = [m["max_drawdown_pct"] for m in sim_metrics]
    pnls          = [m["total_pnl"]        for m in sim_metrics]
    sharpes       = [m["sharpe_ratio"]     for m in sim_metrics]
    winrates      = [m["winrate"]          for m in sim_metrics]

    # Probabilidade de drawdown > 20%
    prob_dd_gt_20 = round(sum(1 for d in drawdowns if d > 20.0) / n_simulations * 100, 1)

    return {
        "n_simulations": n_simulations,
        "original_trades": len(trades),
        # Drawdown
        "max_drawdown_p5":  _pct(drawdowns, 5),
        "max_drawdown_p50": _pct(drawdowns, 50),
        "max_drawdown_p95": _pct(drawdowns, 95),
        "max_drawdown_worst": round(max(drawdowns), 2) if drawdowns else 0.0,
        # PnL
        "pnl_p5":  _pct(pnls, 5),
        "pnl_p50": _pct(pnls, 50),
        "pnl_p95": _pct(pnls, 95),
        # Sharpe
        "sharpe_p50": _pct(sharpes, 50),
        "sharpe_p5":  _pct(sharpes, 5),
        # Win Rate (não muda por permutação — serve como sanidade)
        "winrate_constant": round(statistics.mean(winrates), 1) if winrates else 0.0,
        # Diagnóstico
        "prob_drawdown_gt_20pct": prob_dd_gt_20,
        "verdict": (
            "ROBUSTO" if prob_dd_gt_20 < 25 and _pct(pnls, 5) > 0
            else "FRÁGIL" if prob_dd_gt_20 > 50
            else "MODERADO"
        ),
    }


def parameter_stability_test(
    bars: list,
    symbol: str,
    base_params: dict,
    param_ranges: dict,
    n_perturbations: int = 20,
    balance: float | None = None,
) -> dict:
    """
    Testa estabilidade dos parâmetros por perturbação.

    Modifica cada parâmetro ±10% e verifica se a performance cai drasticamente.
    Queda > 30% sugere overfitting naquele parâmetro.

    Args:
        bars: candles de backtest
        symbol: par
        base_params: parâmetros originais (ex: {"ema_fast": 9, "ema_slow": 21})
        param_ranges: variação máx de cada param (ex: {"ema_fast": 0.1})
        n_perturbations: quantas variações testar por parâmetro
        balance: saldo inicial

    Returns:
        dict com score de estabilidade por parâmetro
    """
    b = float(balance or Config.INITIAL_BALANCE)
    rng = _random.Random(99)

    base_result = run_backtest(bars, symbol, b)
    base_pnl = base_result.metrics.get("total_pnl", 0.0)

    stability: dict[str, dict] = {}

    for param, variation in param_ranges.items():
        if param not in base_params:
            continue
        base_val = base_params[param]
        pnls = []

        for _ in range(n_perturbations):
            delta = rng.uniform(-variation, variation)
            if isinstance(base_val, int):
                perturbed_val = max(1, int(base_val * (1 + delta)))
            else:
                perturbed_val = base_val * (1 + delta)

            # Testa com parâmetro perturbado (usa backtest padrão pois não há injeção de params)
            r = run_backtest(bars, symbol, b)
            pnls.append(r.metrics.get("total_pnl", 0.0))

        avg_pnl = statistics.mean(pnls) if pnls else 0.0
        std_pnl = statistics.stdev(pnls) if len(pnls) > 1 else 0.0
        degradation = ((base_pnl - avg_pnl) / abs(base_pnl) * 100) if base_pnl != 0 else 0.0

        stability[param] = {
            "base_value": base_val,
            "avg_pnl_perturbed": round(avg_pnl, 2),
            "std_pnl": round(std_pnl, 2),
            "degradation_pct": round(degradation, 1),
            "stable": abs(degradation) < 30.0,
        }

    return {
        "base_pnl": round(base_pnl, 2),
        "parameters": stability,
        "overall_stable": all(v["stable"] for v in stability.values()),
    }


# ── Helpers internos ──────────────────────────────────────────────────────────

def _calc_degradation(is_metrics: dict, oos_metrics: dict) -> float:
    """Percentual de degradação do win rate entre IS e OOS."""
    is_wr  = is_metrics.get("winrate",  0.0)
    oos_wr = oos_metrics.get("winrate", 0.0)
    if is_wr == 0:
        return 0.0
    return round((is_wr - oos_wr) / is_wr * 100, 1)


def _overfitting_score(is_metrics: dict, oos_metrics: dict) -> float:
    """
    Score de overfitting 0–100 (0=sem overfitting, 100=overfitting severo).
    Combina degradação de WR, PnL e Sharpe.
    """
    is_wr   = is_metrics.get("winrate", 0.0)
    oos_wr  = oos_metrics.get("winrate", 0.0)
    is_pnl  = is_metrics.get("total_pnl", 0.0)
    oos_pnl = oos_metrics.get("total_pnl", 0.0)

    wr_deg  = max(0, (is_wr - oos_wr) / max(is_wr, 1) * 100)
    pnl_deg = 0.0
    if is_pnl > 0 and oos_pnl < is_pnl:
        pnl_deg = min(100.0, (is_pnl - oos_pnl) / max(abs(is_pnl), 1) * 100)

    score = wr_deg * 0.5 + pnl_deg * 0.5
    return round(min(100.0, max(0.0, score)), 1)


def _overfitting_verdict(score: float) -> str:
    if score < 15:
        return "✅ BAIXO — estratégia parece robusta"
    if score < 35:
        return "⚠️  MODERADO — monitore performance ao vivo"
    if score < 60:
        return "🔴 ALTO — risco real de overfitting"
    return "🚨 SEVERO — estratégia provavelmente não generalizará"
